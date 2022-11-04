def asegurar(func):
    def interna(name):
        groserias = ['Poto', 'Sonzo', 'Fea', 'Feo', 'Tonta', 'Tonto']
        if name in groserias:
            print('No se permiten groserías')
            return
        func(name)
        
    return interna


@asegurar
def saludar(nombre):
    html = open('saludo.html', 'w')
    html.write(f'''
    <!DOCTYPE html>
    <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta http-equiv="X-UA-Compatible" content="IE=edge">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Document</title>
        </head>
        <body>
            <h3>Hola {nombre}</h3>
        </body>
    </html>
    ''')
    html.close()

@asegurar
def beetlejuice(nombre):
    print(nombre)
    print(nombre)
    print(nombre)

@asegurar
def excel(nombre):
    from openpyxl import Workbook
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = 'Hola'
    ws['A2'] = nombre
    ws['A3'] = 'Mucho gusto!'


    # Python types will automatically be converted
    import datetime
    ws['B2'] = datetime.datetime.now()

    # Save the file
    wb.save("saludo.xlsx")

# saludar('Poto')
# beetlejuice('Feo')
excel('Yojan')